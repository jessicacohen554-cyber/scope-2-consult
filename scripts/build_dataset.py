#!/usr/bin/env python3
"""Reshape the GHG Protocol Scope 2 public consultation raw export into a tidy,
queryable dataset.

Input : S2-PublicConsultationFeedback-RawData-2026.07.29.xlsx  (1 sheet, 1072 x 181)
Output: data/*.csv, data/scope2_consultation.sqlite, data/scope2_consultation.xlsx

Every answer stays attached to its respondent ID; nothing is dropped or
truncated. Interpretation (sections, scale directions, ordinal ladders) lives in
survey_meta.py.

Usage:  python3 scripts/build_dataset.py
"""
from __future__ import annotations

import csv
import json
import re
import sqlite3
import sys
import unicodedata
from collections import Counter, defaultdict
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parent))
import survey_meta as meta  # noqa: E402

ROOT = Path(__file__).resolve().parent.parent
SRC = ROOT / "S2-PublicConsultationFeedback-RawData-2026.07.29.xlsx"
OUT = ROOT / "data"
SHEET = "clean_data_final1"

# A cell exactly this long was almost certainly cut off by the survey tool.
TEXT_LIMIT = 4000

# Multi-select cells are semicolon-delimited with a trailing delimiter. A part is
# treated as one of the offered options if enough respondents picked it; the rest
# are retained as write-ins so no text is lost.
CANONICAL_MIN_SHARE = 0.05
OTHER_PAT = re.compile(r"^other\b|\(please (specify|explain|describe)\)|\(describe\)|\(explain\)", re.I)


# ---------------------------------------------------------------------------
# text normalisation
# ---------------------------------------------------------------------------
def norm_text(value) -> str:
    """Collapse the export's assorted invisible whitespace without changing meaning.

    The raw file carries NBSP, narrow NBSP, thin space, ideographic space, BOM,
    zero-width space and zero-width joiner - a side effect of respondents pasting
    from Word and of the survey tool's own encoding. Left alone they split
    otherwise-identical option labels into distinct strings.
    """
    if value is None:
        return ""
    s = str(value)
    out = []
    for ch in s:
        cat = unicodedata.category(ch)
        if cat == "Cf":  # zero-width / formatting: drop entirely
            continue
        if cat == "Zs" and ch != " ":  # any other space separator -> plain space
            out.append(" ")
            continue
        out.append(ch)
    s = "".join(out)
    s = s.replace("\r\n", "\n").replace("\r", "\n")
    s = re.sub(r"[ \t]+", " ", s)
    s = re.sub(r"\n{3,}", "\n\n", s)
    s = re.sub(r" *\n *", "\n", s)
    # A couple of cells were stored as JSON-ish literals: ["4 - Mostly ready"]
    m = re.fullmatch(r'\["(.*)"\]', s.strip())
    if m:
        s = m.group(1)
    return s.strip()


def norm_key(s: str) -> str:
    """Loose key for matching option labels against the ladders in survey_meta."""
    s = norm_text(s).lower()
    s = s.replace("–", "-").replace("—", "-").replace("−", "-")
    s = s.replace("’", "'").replace("‘", "'")
    return re.sub(r"\s+", " ", s).strip(" .")


# ---------------------------------------------------------------------------
# question metadata derivation
# ---------------------------------------------------------------------------
ROLE_RULES = [
    ("other_specify", re.compile(
        r"if you (selected|answered)\s*[\"'“]?other|^if you selected .other", re.I)),
    ("reasons_support", re.compile(r"reasons?\s+(for|of)\s+support", re.I)),
    ("reasons_oppose", re.compile(
        r"why you are not support|not supportive|reasons?\s+(for|of)\s+concern|your concerns", re.I)),
    ("basis", re.compile(r"basis for your assessment|basis for the assessment", re.I)),
    ("comment", re.compile(
        r"please provide (any )?(additional )?(comments|context|explanations|further details|"
        r"additional feedback)|additional comments|additional context|please briefly explain|"
        r"please explain why|please describe|please provide details", re.I)),
]

REF_PAT = re.compile(r"question[s]?\s*(\d{1,3})(?:\s*(?:[-–&,]|and|to)\s*(\d{1,3}))?", re.I)
Q_SHORT_PAT = re.compile(r"\bQ\.?\s?(\d{1,3})\b")


def classify_role(qnum: int, text: str, is_multi: bool, is_free: bool) -> str:
    if 3 <= qnum <= 17:
        return "profile"
    for role, pat in ROLE_RULES:
        if pat.search(text):
            if role in ("reasons_support", "reasons_oppose"):
                # "Please provide comments regarding your reasons for support" is a
                # comment on the reasons, not the reasons picklist itself.
                if re.search(r"^\s*\d+\.\s*please provide (any )?(additional )?comments", text, re.I):
                    return "comment"
                return role if is_multi else "comment"
            return role
    return "free_text_primary" if is_free else "primary"


def extract_refs(text: str, self_num: int) -> list[int]:
    refs: set[int] = set()
    for m in REF_PAT.finditer(text):
        a = int(m.group(1))
        b = int(m.group(2)) if m.group(2) else None
        if b and 0 < b - a <= 12:
            refs.update(range(a, b + 1))
        else:
            refs.add(a)
            if b:
                refs.add(b)
    for m in Q_SHORT_PAT.finditer(text):
        refs.add(int(m.group(1)))
    refs.discard(self_num)
    return sorted(r for r in refs if 3 <= r <= 183)


# ---------------------------------------------------------------------------
# load
# ---------------------------------------------------------------------------
def load_grid():
    wb = openpyxl.load_workbook(SRC, read_only=True, data_only=True)
    ws = wb[SHEET]
    it = ws.iter_rows(values_only=True)
    header = [norm_text(h) for h in next(it)]
    rows = []
    for r in it:
        vals = [norm_text(v) for v in r]
        if any(vals):
            rows.append(vals)
    wb.close()
    return header, rows


def main() -> None:
    OUT.mkdir(exist_ok=True)
    header, rows = load_grid()
    n_resp = len(rows)
    print(f"loaded {n_resp} respondents x {len(header)} columns")

    # ---------------- pass 1: shape of each column ----------------
    cols = []  # dicts, one per column
    for idx, h in enumerate(header):
        vals = [r[idx] for r in rows]
        filled = [v for v in vals if v]
        m = re.match(r"\s*(\d{1,3})\s*\.", h)
        qnum = int(m.group(1)) if m else None

        ends_semi = sum(1 for v in filled if v.endswith(";"))
        is_multi = bool(filled) and ends_semi >= 0.9 * len(filled)

        parts = Counter()
        if is_multi:
            for v in filled:
                for p in v.split(";"):
                    p = p.strip()
                    if p:
                        parts[p] += 1
        whole = Counter(filled)
        maxlen = max((len(v) for v in filled), default=0)
        # Single-select: no delimiter, all answers short, and answers repeat.
        # No free-text column in this export has a longest answer under 320 chars,
        # so the length test does most of the work; the repeat test additionally
        # keeps Q4 (name) and Q5 (organisation) out, and lets in the larger
        # picklists - Q6 country (56), Q16 sector (29), Q14 organisation type (15).
        is_single = (
            (not is_multi) and bool(filled) and maxlen <= 320
            and (len(whole) <= 12 or len(whole) / len(filled) <= 0.35)
        )
        is_free = not (is_multi or is_single) and qnum is not None

        cols.append(dict(
            col_index=idx + 1, header=h, qnum=qnum, values=vals, filled=filled,
            is_multi=is_multi, is_single=is_single, is_free=is_free,
            parts=parts, whole=whole, maxlen=maxlen,
        ))

    id_col = cols[0]
    assert id_col["header"] == "ID", id_col["header"]
    resp_ids = [int(v) for v in id_col["values"]]
    assert len(set(resp_ids)) == n_resp, "respondent IDs are not unique"

    qcols = [c for c in cols if c["qnum"] is not None]
    by_qnum = {c["qnum"]: c for c in qcols}

    # ---------------- hand-curated question labels ----------------
    # Left join by question_id, but a miss on either side is an error rather
    # than a null: an unlabelled question would surface as blank cells in
    # questions.csv long after the fact, and a label for a question that is not
    # in the export means the file has drifted from the source. The vocabulary
    # and its rules live in survey_meta; this is only the join.
    labels = meta.load_labels(ROOT)
    qids = {f"Q{c['qnum']:03d}" for c in qcols}
    missing, extra = sorted(qids - set(labels)), sorted(set(labels) - qids)
    if missing or extra:
        raise SystemExit(
            f"{meta.LABELS_FILE}: label coverage mismatch\n"
            + (f"  unlabelled questions ({len(missing)}): "
               f"{', '.join(missing)}\n" if missing else "")
            + (f"  labels for questions absent from the export ({len(extra)}): "
               f"{', '.join(extra)}\n" if extra else ""))
    print(f"joined {len(labels)} question labels")

    # ---------------- pass 2: question type + interpretation ----------------
    for c in qcols:
        qnum, h = c["qnum"], c["header"]
        c["question_id"] = f"Q{qnum:03d}"
        c["text"] = re.sub(r"^\s*\d{1,3}\s*\.\s*", "", h).strip()
        order, label = meta.section_for(qnum)
        c["section_order"], c["section"] = order, label
        c["role"] = classify_role(qnum, h, c["is_multi"], c["is_free"])
        c["refs"] = extract_refs(h, qnum)

        scale = meta.SCALES.get(qnum)
        ladder = meta.ORDINAL_LADDERS.get(qnum)
        numeric_only = c["is_single"] and all(
            re.fullmatch(r"[1-5]", v) for v in c["filled"])

        if c["is_multi"]:
            c["qtype"] = "multi_select"
        elif scale:
            c["qtype"] = "likert_1_5" if numeric_only else "scale_labeled"
        elif qnum == 183:
            c["qtype"] = "numeric_year"
        elif ladder:
            c["qtype"] = "ordinal_select"
        elif c["is_single"]:
            c["qtype"] = "single_select"
        else:
            c["qtype"] = "free_text"

        c["construct"] = scale[0] if scale else ("ordinal" if ladder else "")
        c["anchor_low"] = scale[1] if scale else ""
        c["anchor_high"] = scale[2] if scale else ""
        c["scale_note"] = scale[3] if scale else ""
        c["note"] = meta.NOTES.get(qnum, "")
        c["ladder"] = ladder
        c["labels"] = labels[c["question_id"]]

    # parent / anchor: what a follow-on question hangs off.
    ordered = sorted(qcols, key=lambda c: c["qnum"])
    FOLLOW_ON = {"other_specify", "reasons_support", "reasons_oppose", "comment", "basis"}
    for i, c in enumerate(ordered):
        parent = anchor = None
        if c["role"] in FOLLOW_ON:
            explicit = [r for r in c["refs"] if r < c["qnum"] and r in by_qnum]
            parent = max(explicit) if explicit else (
                ordered[i - 1]["qnum"] if i > 0 else None)
            for prev in reversed(ordered[:i]):
                if prev["role"] in ("primary", "profile") and prev["section_order"] == c["section_order"]:
                    anchor = prev["qnum"]
                    break
        elif c["role"] in ("primary", "free_text_primary"):
            anchor = c["qnum"]
        c["parent_qnum"], c["anchor_qnum"] = parent, anchor

    # ---------------- pass 3: option catalogue ----------------
    option_rows = []
    for c in qcols:
        if c["is_multi"]:
            n = len(c["filled"])
            cutoff = max(3, CANONICAL_MIN_SHARE * n)
            for opt, cnt in c["parts"].most_common():
                is_other = bool(OTHER_PAT.search(opt))
                canonical = (
                    cnt >= cutoff
                    or (is_other and len(opt) <= 60)
                    or norm_key(opt) in meta.STANDARD_OPTIONS
                )
                option_rows.append(dict(
                    question_id=c["question_id"], question_number=c["qnum"],
                    option_text=opt, option_rank="",
                    is_canonical=int(canonical), is_other_option=int(is_other),
                    is_off_scale=0, n_selected=cnt,
                    pct_of_answered=round(100.0 * cnt / n, 2) if n else 0.0,
                ))
        elif c["qtype"] in ("single_select", "ordinal_select", "scale_labeled", "likert_1_5"):
            n = len(c["filled"])
            for opt, cnt in c["whole"].most_common():
                key = norm_key(opt)
                rank = ""
                if c["ladder"] and key in c["ladder"]:
                    rank = c["ladder"][key]
                elif c["qtype"] in ("likert_1_5", "scale_labeled"):
                    m = re.match(r"([1-5])\b", opt)
                    if m and key not in meta.OFF_SCALE:
                        rank = int(m.group(1))
                option_rows.append(dict(
                    question_id=c["question_id"], question_number=c["qnum"],
                    option_text=opt, option_rank=rank,
                    is_canonical=1, is_other_option=int(bool(OTHER_PAT.search(opt))),
                    is_off_scale=int(key in meta.OFF_SCALE), n_selected=cnt,
                    pct_of_answered=round(100.0 * cnt / n, 2) if n else 0.0,
                ))
    canonical_opts = {
        (o["question_id"], o["option_text"]) for o in option_rows if o["is_canonical"]
    }
    # Derived rather than hand-maintained, so the flags cannot drift from the data.
    write_in_counts = Counter()
    borderline = Counter()
    for o in option_rows:
        if not o["is_canonical"]:
            write_in_counts[o["question_number"]] += o["n_selected"]
            borderline[o["question_number"]] = max(
                borderline[o["question_number"]], o["n_selected"])
    rank_lookup = {
        (o["question_id"], o["option_text"]): o["option_rank"] for o in option_rows
    }

    # ---------------- pass 4: long responses + exploded selections ----------
    long_rows, sel_rows = [], []
    per_resp = defaultdict(lambda: dict(answered=0, free_text=0, chars=0, selections=0))

    for ri, rid in enumerate(resp_ids):
        for c in qcols:
            raw = c["values"][ri]
            if not raw:
                continue
            qid, qnum = c["question_id"], c["qnum"]
            stats = per_resp[rid]
            stats["answered"] += 1

            numeric = ""
            n_sel = ""
            if c["is_multi"]:
                picked = [p.strip() for p in raw.split(";") if p.strip()]
                n_sel = len(picked)
                stats["selections"] += n_sel
                for si, opt in enumerate(picked, 1):
                    sel_rows.append(dict(
                        respondent_id=rid, question_id=qid, question_number=qnum,
                        selection_index=si, option_text=opt,
                        is_canonical=int((qid, opt) in canonical_opts),
                    ))
            elif c["qtype"] == "numeric_year":
                m = re.fullmatch(r"(\d{4})", raw)
                numeric = int(m.group(1)) if m else ""
            else:
                r = rank_lookup.get((qid, raw), "")
                numeric = r if r != "" else ""

            is_text = c["qtype"] in ("free_text",)
            if is_text:
                stats["free_text"] += 1
                stats["chars"] += len(raw)

            lab = c["labels"]
            long_rows.append(dict(
                respondent_id=rid, question_id=qid, question_number=qnum,
                section_order=c["section_order"], section=c["section"],
                # Carried at answer grain, not just on the codebook, so that a
                # discrete answer can be grouped by shorthand, method, concern
                # or answer type without joining back to questions.
                shorthand=lab["shorthand"], method=lab["method"],
                category=lab["category"], policy_lever=lab["policy_lever"],
                asks_for=lab["asks_for"],
                question_type=c["qtype"], role=c["role"], construct=c["construct"],
                answer_text=raw, answer_numeric=numeric, n_selected=n_sel,
                char_count=len(raw) if is_text else "",
                word_count=len(raw.split()) if is_text else "",
                likely_truncated=int(is_text and len(raw) >= TEXT_LIMIT),
            ))

    # ---------------- respondents ----------------
    PROFILE = {
        3: "redaction_requested", 4: "name", 5: "organization", 6: "country",
        8: "newsletter_optin", 9: "responding_as", 10: "has_ghg_inventory",
        11: "has_ghg_inventory_other", 12: "involved_in_inventory",
        13: "involved_in_inventory_other", 14: "organization_type",
        15: "organization_type_other", 16: "sector", 17: "sector_other",
    }
    n_answerable = len(qcols) - len(PROFILE)

    def attribution(row):
        """Label withheld identities instead of leaving them null.

        Requesting redaction is itself a finding worth counting, so the two
        reasons a name can be missing are kept distinct: 'Redacted' means the
        respondent asked for redaction (question 3), 'Not provided' means they
        did not ask but left the field empty anyway.
        """
        redacted = row["redaction_requested"] == "Yes"
        for field in ("name", "organization"):
            if not row[field]:
                row[field] = "Redacted" if redacted else "Not provided"
        return int(redacted)

    resp_rows = []
    for ri, rid in enumerate(resp_ids):
        row = {"respondent_id": rid}
        for qnum, field in PROFILE.items():
            row[field] = by_qnum[qnum]["values"][ri] if qnum in by_qnum else ""
        s = per_resp[rid]
        substantive = s["answered"] - sum(
            1 for qnum in PROFILE if qnum in by_qnum and by_qnum[qnum]["values"][ri])
        row["is_redacted"] = attribution(row)
        row.update(
            n_questions_answered=s["answered"],
            n_substantive_answered=substantive,
            substantive_completion_pct=round(100.0 * substantive / n_answerable, 1),
            n_free_text_answers=s["free_text"],
            free_text_chars=s["chars"],
            n_multi_select_selections=s["selections"],
        )
        resp_rows.append(row)

    # ---------------- questions codebook ----------------
    q_rows = []
    for c in ordered:
        n = len(c["filled"])
        n_opts = len([o for o in option_rows
                      if o["question_id"] == c["question_id"] and o["is_canonical"]])
        lab = c["labels"]
        q_rows.append(dict(
            question_id=c["question_id"], question_number=c["qnum"],
            column_index=c["col_index"],
            section_order=c["section_order"], section=c["section"],
            shorthand=lab["shorthand"], label=lab["label"],
            method=lab["method"], category=lab["category"],
            subcategory=lab["subcategory"], policy_lever=lab["policy_lever"],
            asks_for=lab["asks_for"],
            question_type=c["qtype"], role=c["role"],
            scale_construct=c["construct"],
            scale_anchor_low=c["anchor_low"], scale_anchor_high=c["anchor_high"],
            scale_note=c["scale_note"],
            parent_question=c["parent_qnum"] or "", anchor_question=c["anchor_qnum"] or "",
            references_questions=",".join(str(r) for r in c["refs"]),
            n_answered=n, response_rate_pct=round(100.0 * n / n_resp, 1),
            n_options=n_opts if n_opts else "",
            n_write_in_selections=write_in_counts.get(c["qnum"], 0),
            # Set where a string classified as a write-in was nonetheless picked by
            # 5+ respondents, i.e. the offered-option call is worth eyeballing.
            review_option_split=int(borderline.get(c["qnum"], 0) >= 5),
            max_answer_chars=c["maxlen"],
            question_text=c["text"], question_text_short=c["text"][:120],
            notes=c["note"], label_notes=lab["label_notes"],
        ))

    # ---------------- wide table ----------------
    wide_fields = (["respondent_id"]
                   + [PROFILE[q] for q in sorted(PROFILE) if q in by_qnum]
                   + ["is_redacted"])
    numeric_qs = [c for c in ordered
                  if c["qtype"] in ("likert_1_5", "scale_labeled", "ordinal_select", "numeric_year")]
    body_qs = [c for c in ordered if c["qnum"] not in PROFILE]
    for c in body_qs:
        wide_fields.append(c["question_id"])
        if c in numeric_qs:
            wide_fields.append(c["question_id"] + "_num")
    wide_rows = []
    for ri, rid in enumerate(resp_ids):
        row = {"respondent_id": rid}
        for qnum in sorted(PROFILE):
            if qnum in by_qnum:
                row[PROFILE[qnum]] = by_qnum[qnum]["values"][ri]
        row["is_redacted"] = attribution(row)
        for c in body_qs:
            raw = c["values"][ri]
            row[c["question_id"]] = raw
            if c in numeric_qs:
                if c["qtype"] == "numeric_year":
                    m = re.fullmatch(r"(\d{4})", raw)
                    row[c["question_id"] + "_num"] = int(m.group(1)) if m else ""
                else:
                    row[c["question_id"] + "_num"] = rank_lookup.get(
                        (c["question_id"], raw), "") if raw else ""
        wide_rows.append(row)

    # ---------------- write CSVs ----------------
    def write_csv(name, rows_, fields):
        path = OUT / name
        with path.open("w", newline="", encoding="utf-8") as f:
            w = csv.DictWriter(f, fieldnames=fields, extrasaction="raise")
            w.writeheader()
            w.writerows(rows_)
        print(f"  {name}: {len(rows_)} rows x {len(fields)} cols")

    print("writing CSVs")
    write_csv("questions.csv", q_rows, list(q_rows[0]))
    write_csv("question_options.csv", option_rows, list(option_rows[0]))
    write_csv("respondents.csv", resp_rows, list(resp_rows[0]))
    write_csv("responses_long.csv", long_rows, list(long_rows[0]))
    write_csv("response_selections.csv", sel_rows, list(sel_rows[0]))
    write_csv("responses_wide.csv", wide_rows, wide_fields)

    # ---------------- SQLite ----------------
    db_path = OUT / "scope2_consultation.sqlite"
    if db_path.exists():
        db_path.unlink()
    con = sqlite3.connect(db_path)
    con.executescript("""
    CREATE TABLE questions (
      question_id TEXT PRIMARY KEY, question_number INTEGER, column_index INTEGER,
      section_order INTEGER, section TEXT,
      shorthand TEXT, label TEXT, method TEXT, category TEXT, subcategory TEXT,
      policy_lever TEXT, asks_for TEXT,
      question_type TEXT, role TEXT,
      scale_construct TEXT, scale_anchor_low TEXT, scale_anchor_high TEXT,
      scale_note TEXT, parent_question INTEGER, anchor_question INTEGER,
      references_questions TEXT, n_answered INTEGER, response_rate_pct REAL,
      n_options INTEGER, n_write_in_selections INTEGER, review_option_split INTEGER,
      max_answer_chars INTEGER,
      question_text TEXT, question_text_short TEXT, notes TEXT,
      label_notes TEXT);
    CREATE TABLE question_options (
      question_id TEXT, question_number INTEGER, option_text TEXT,
      option_rank INTEGER, is_canonical INTEGER, is_other_option INTEGER,
      is_off_scale INTEGER, n_selected INTEGER, pct_of_answered REAL);
    CREATE TABLE respondents (respondent_id INTEGER PRIMARY KEY);
    CREATE TABLE responses (
      respondent_id INTEGER, question_id TEXT, question_number INTEGER,
      section_order INTEGER, section TEXT,
      shorthand TEXT, method TEXT, category TEXT, policy_lever TEXT,
      asks_for TEXT,
      question_type TEXT, role TEXT,
      construct TEXT, answer_text TEXT, answer_numeric REAL, n_selected INTEGER,
      char_count INTEGER, word_count INTEGER, likely_truncated INTEGER);
    CREATE TABLE response_selections (
      respondent_id INTEGER, question_id TEXT, question_number INTEGER,
      selection_index INTEGER, option_text TEXT, is_canonical INTEGER);
    """)
    # respondents table: build columns from the dict keys
    rfields = list(resp_rows[0])
    con.execute("DROP TABLE respondents")
    con.execute("CREATE TABLE respondents (%s)" % ", ".join(
        f'"{k}" {"INTEGER" if k == "respondent_id" or k.startswith("n_") or k.endswith("_chars") else "TEXT"}'
        for k in rfields))

    def insert(table, rows_, fields):
        con.executemany(
            f'INSERT INTO {table} ({",".join(chr(34)+f+chr(34) for f in fields)}) '
            f'VALUES ({",".join("?" * len(fields))})',
            [tuple(None if r.get(f) == "" else r.get(f) for f in fields) for r in rows_])

    insert("questions", q_rows, list(q_rows[0]))
    insert("question_options", option_rows, list(option_rows[0]))
    insert("respondents", resp_rows, rfields)
    insert("responses", long_rows, list(long_rows[0]))
    insert("response_selections", sel_rows, list(sel_rows[0]))

    con.executescript("""
    CREATE INDEX ix_resp_q   ON responses(question_number);
    CREATE INDEX ix_resp_r   ON responses(respondent_id);
    CREATE INDEX ix_resp_sec ON responses(section_order);
    CREATE INDEX ix_resp_sh  ON responses(shorthand);
    CREATE INDEX ix_resp_ask ON responses(asks_for);
    CREATE INDEX ix_sel_q    ON response_selections(question_number);
    CREATE INDEX ix_sel_r    ON response_selections(respondent_id);

    -- Every scale answer with its respondent's profile and the scale's meaning.
    -- One row per respondent x question: the labels are here to group discrete
    -- answers by, not to aggregate them away.
    CREATE VIEW v_scale_answers AS
    SELECT r.respondent_id, r.question_number, q.question_id,
           q.shorthand, q.label, q.method, q.category, q.policy_lever,
           q.asks_for, q.section,
           q.scale_construct, q.scale_anchor_low, q.scale_anchor_high,
           q.question_text_short, r.answer_text, r.answer_numeric,
           p.country, p.responding_as, p.organization_type, p.sector,
           p.organization, p.is_redacted
    FROM responses r
    JOIN questions q USING(question_id)
    JOIN respondents p ON p.respondent_id = r.respondent_id
    WHERE q.question_type IN ('likert_1_5','scale_labeled')
      AND r.answer_numeric IS NOT NULL;

    -- Mean/median support etc. per question, construct kept visible so that
    -- support and burden scales are never silently pooled. shorthand is the
    -- column to read: question_text_short buries the subject.
    CREATE VIEW v_scale_summary AS
    SELECT q.question_id, q.question_number,
           q.shorthand, q.label, q.method, q.category, q.subcategory,
           q.policy_lever, q.asks_for, q.section, q.scale_construct,
           q.scale_anchor_low, q.scale_anchor_high, q.question_text_short,
           COUNT(r.answer_numeric) AS n_scored,
           ROUND(AVG(r.answer_numeric), 2) AS mean_score,
           SUM(r.answer_numeric <= 2) AS n_low_1_2,
           SUM(r.answer_numeric = 3)  AS n_mid_3,
           SUM(r.answer_numeric >= 4) AS n_high_4_5
    FROM questions q LEFT JOIN responses r USING(question_id)
    WHERE q.question_type IN ('likert_1_5','scale_labeled')
    GROUP BY q.question_id ORDER BY q.question_number;

    -- Option frequencies for choice questions, with profile joins available.
    -- One row per option ticked per respondent.
    CREATE VIEW v_selections AS
    SELECT s.respondent_id, s.question_number, s.question_id,
           q.shorthand, q.label, q.method, q.category, q.policy_lever,
           q.asks_for, q.section,
           q.question_text_short, s.option_text, s.is_canonical,
           p.country, p.responding_as, p.organization_type, p.sector,
           p.is_redacted
    FROM response_selections s
    JOIN questions q USING(question_id)
    JOIN respondents p ON p.respondent_id = s.respondent_id;

    -- Drill-down tree: a substantive question with the follow-ups hanging off
    -- it. Shorthands on both sides, so the 62 boilerplate follow-ups are
    -- readable without their parent's wording.
    CREATE VIEW v_question_tree AS
    SELECT a.question_number AS anchor_number, a.shorthand AS anchor_shorthand,
           a.label AS anchor_label, a.question_text_short AS anchor_text,
           a.method, a.category, a.policy_lever, a.section,
           c.question_number, c.question_id, c.shorthand, c.label, c.asks_for,
           c.role, c.question_type, c.n_answered, c.question_text_short
    FROM questions c JOIN questions a ON a.question_number = c.anchor_question
    ORDER BY a.question_number, c.question_number;

    -- How many answers of each type each respondent gave. Counts only; the
    -- discrete answers stay in `responses`, which this groups rather than
    -- replaces, so any cell here drills back to named respondents.
    CREATE VIEW v_answer_types AS
    SELECT r.respondent_id, q.method, q.category, q.asks_for, q.question_type,
           COUNT(*)                             AS n_answers,
           COUNT(r.answer_numeric)              AS n_scored,
           SUM(COALESCE(r.n_selected, 0))       AS n_selections,
           SUM(COALESCE(r.char_count, 0))       AS free_text_chars
    FROM responses r JOIN questions q USING(question_id)
    GROUP BY r.respondent_id, q.method, q.category, q.asks_for, q.question_type;

    -- Every distinct answer to a choice question with how many respondents gave
    -- it, keyed by shorthand rather than by wording.
    CREATE VIEW v_option_counts AS
    SELECT q.question_number, q.question_id, q.shorthand, q.label,
           q.method, q.category, q.policy_lever, q.asks_for, q.question_type,
           o.option_text, o.option_rank, o.is_canonical, o.is_other_option,
           o.is_off_scale, o.n_selected, o.pct_of_answered, q.n_answered
    FROM question_options o JOIN questions q USING(question_id)
    ORDER BY q.question_number, o.n_selected DESC;

    -- Redaction as a finding in its own right: does withholding identity travel
    -- with who the respondent is, how much they wrote, or how they answered?
    CREATE VIEW v_redaction_profile AS
    SELECT is_redacted, responding_as, organization_type,
           COUNT(*) AS n_respondents,
           ROUND(AVG(substantive_completion_pct), 1) AS mean_completion_pct,
           ROUND(AVG(free_text_chars)) AS mean_free_text_chars,
           ROUND(AVG(n_free_text_answers), 1) AS mean_free_text_answers
    FROM respondents
    GROUP BY is_redacted, responding_as, organization_type;

    -- Mean score per scale question, split by whether identity was withheld.
    CREATE VIEW v_scale_by_redaction AS
    SELECT question_number, shorthand, label, method, policy_lever,
           question_text_short, scale_construct,
           SUM(is_redacted = 0) AS n_named,
           ROUND(AVG(CASE WHEN is_redacted = 0 THEN answer_numeric END), 2) AS mean_named,
           SUM(is_redacted = 1) AS n_redacted,
           ROUND(AVG(CASE WHEN is_redacted = 1 THEN answer_numeric END), 2) AS mean_redacted
    FROM v_scale_answers
    GROUP BY question_number ORDER BY question_number;

    -- All free text for a respondent, in survey order.
    CREATE VIEW v_free_text AS
    SELECT r.respondent_id, r.question_number, q.question_id,
           q.shorthand, q.label, q.method, q.category, q.policy_lever,
           q.asks_for, q.section, q.role,
           q.question_text_short, r.char_count, r.word_count, r.likely_truncated,
           r.answer_text
    FROM responses r JOIN questions q USING(question_id)
    WHERE q.question_type = 'free_text';
    """)
    con.commit()
    con.close()
    print(f"  scope2_consultation.sqlite: {db_path.stat().st_size/1e6:.1f} MB")

    # ---------------- Excel workbook ----------------
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        wb.remove(wb.active)
        # The long tables are deliberately left out of the workbook: they restate
        # the same free text a third time (~19 MB of the file) and long format is
        # awkward in Excel anyway. responses_wide holds every answer already, one
        # row per respondent, which is the shape PivotTables want. Use the CSVs or
        # the SQLite database for long-format work.
        sheets = [
            ("questions", q_rows, list(q_rows[0])),
            ("question_options", option_rows, list(option_rows[0])),
            ("respondents", resp_rows, rfields),
            ("responses_wide", wide_rows, wide_fields),
        ]
        for name, rws, flds in sheets:
            ws = wb.create_sheet(name)
            ws.append(flds)
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(vertical="top")
            for r in rws:
                ws.append([r.get(f, "") if r.get(f, "") != "" else None for f in flds])
            ws.freeze_panes = "B2"
            ws.auto_filter.ref = (
                f"A1:{get_column_letter(len(flds))}{len(rws)+1}")
            for i, f in enumerate(flds, 1):
                ws.column_dimensions[get_column_letter(i)].width = min(
                    46, max(11, len(str(f)) + 2))
        xlsx_path = OUT / "scope2_consultation.xlsx"
        wb.save(xlsx_path)
        print(f"  scope2_consultation.xlsx: {xlsx_path.stat().st_size/1e6:.1f} MB")
    except Exception as exc:  # pragma: no cover
        print(f"  (workbook skipped: {exc})")

    # ---------------- manifest ----------------
    manifest = dict(
        source_file=SRC.name, sheet=SHEET,
        respondents=n_resp, questions=len(qcols), raw_columns=len(header),
        question_numbers_absent_from_export=[
            q for q in range(1, 184) if q not in by_qnum],
        rows=dict(responses_long=len(long_rows), response_selections=len(sel_rows),
                  question_options=len(option_rows)),
        question_types=dict(Counter(c["qtype"] for c in ordered)),
        roles=dict(Counter(c["role"] for c in ordered)),
        labelled_questions=len(labels),
        methods=dict(Counter(c["labels"]["method"] for c in ordered)),
        categories=dict(Counter(c["labels"]["category"] for c in ordered)),
        asks_for=dict(Counter(c["labels"]["asks_for"] for c in ordered)),
        policy_levers=len({c["labels"]["policy_lever"] for c in ordered}),
        subcategories=len({c["labels"]["subcategory"] for c in ordered}),
        scale_constructs=dict(Counter(c["construct"] for c in ordered if c["construct"])),
        truncated_free_text_answers=sum(r["likely_truncated"] for r in long_rows),
        non_canonical_selections=sum(1 for s in sel_rows if not s["is_canonical"]),
    )
    (OUT / "manifest.json").write_text(json.dumps(manifest, indent=2), encoding="utf-8")
    print("  manifest.json")
    print(json.dumps(manifest, indent=2))


if __name__ == "__main__":
    main()
